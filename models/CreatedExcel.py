import openpyxl as opl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
import os


class ExcelFinal:

    def __init__(self, sheet_names=[], rows=[], columns=[], values=[]):
        self.sheet_names = sheet_names
        self.rows = rows
        self.columns = columns
        self.values = values

    @property
    def get_file_path(self):
        FILE_NAME = "Compared_excel.xlsx"
        dir = os.getcwd()
        save_directory = os.path.join(dir, FILE_NAME)
        return save_directory

    def to_excel(self):
        wb = self.create_workbook()
        self.create_workbook_sheets(wb)
        for i, sheet_name in enumerate(wb.sheetnames):
            sheet = wb.get_sheet_by_name(sheet_name)
            try:
                # print(i, self.values[i])
                self.add_values(sheet, self.values[i])
                self.set_row_colors(self.rows[i], sheet)
            except IndexError:
                continue
        wb.save(self.get_file_path)

    def create_workbook(self):
        workbook = opl.Workbook()
        return workbook

    def create_workbook_sheets(self, workbook: opl.Workbook):
        sheets = self.sheet_names
        for i in range(0, len(sheets)):
            workbook.create_sheet(sheets[i], i)
        workbook.save(self.get_file_path)

    def add_values(self, sheet: Worksheet, values: list[list]):
        for row_values in values:
            sheet.append(row_values)

    def set_row_colors(self, rows: (int, int), sheet: Worksheet):
        row_count = sheet.max_row
        idx, count = rows
        print("ROW_COUNT:", row_count, "COUNT:", count)
        min_row_count = row_count - count
        if row_count == min_row_count:
            print("MIN ROW COUNT IS THE SAME AS ROW COUNT")
            pass
            # Add color to all the sheet value cells
            # return

        sheet_rows = sheet.iter_rows(min_row=min_row_count, max_row=row_count,
                                     min_col=1, max_col=sheet.max_column)
        print(min_row_count, row_count, 1, sheet.max_column)
        for row in sheet_rows:
            for value in row:
                value.fill = PatternFill(bgColor="F5B042")
