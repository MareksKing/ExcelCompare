from dataclasses import dataclass
from .Sheets import Sheet
from openpyxl import load_workbook, Workbook

@dataclass
class Excel:

    file_path: str

    @property
    def workbook(self):
        workbook: Workbook = load_workbook(self.file_path, read_only=True)
        return workbook
        

    @property
    def sheet_names(self):
        sheet_names = self.workbook.get_sheet_names()
        return sheet_names

    @property
    def sheets(self):
        sheets: list[Sheet] = []
        for sheet_name in self.sheet_names:
            sheet = self.workbook.get_sheet_by_name(sheet_name)
            
            sheets.append(Sheet(sheet_name, sheet.max_row, sheet.max_column, sheet.values))

        return sheets

    def describe(self) -> None:
        print(f"{self}")
        print(f"Workbook: {self.workbook}")
        print(f"Sheetnames: {self.sheet_names}")
        print(f"Sheets: {self.sheets}")
